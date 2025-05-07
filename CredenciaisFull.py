# === IMPORTAÇÕES ===
import requests
import subprocess
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from datetime import datetime
import os
import re
import time
import sys

# Tirar os warnings chatos
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# configurações
script_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
arquivo_excel = os.path.join(script_dir, 'Credenciais.xlsx')


def buscar_cpf_por_re(re):
    url = f"https://api.com.br/api/re/{re}/dadosResumidos" 
    try:
        response = requests.get(url, verify=False, timeout=10)
        response.raise_for_status()
        data = response.json()
        if isinstance(data.get('dados'), list) and data['dados']:
            return data['dados'][0].get('cpf', {}).get('cpfComDigito')
    except Exception as e:
        return "NÃO", f"Erro na consulta a API do Swagger: {str(e)}"
    return None

def consultar_dados_por_cpf(cpf):
    resultado = {"nome": "Nome não encontrado", "email_funcional": "Email não encontrado"}
    try:
        url_nome = f"https://api.com.br/api/cpf/{cpf}/dadosResumidos" 
        r_nome = requests.get(url_nome, verify=False, timeout=10)
        r_nome.raise_for_status()
        dados = r_nome.json().get('dados')
        if dados:
            resultado["nome"] = dados[0].get('nomeCompleto', resultado["nome"])
    except Exception as e:
        return "NÃO", f"Erro na consulta a API do Swagger: {str(e)}"

    try:
        url_contato = f"https://api.com.br/api/cpf/{cpf}/informacaoContato" 
        r_contato = requests.get(url_contato, verify=False, timeout=10)
        r_contato.raise_for_status()
        contatos = r_contato.json().get('dados')
        if contatos:
            for email_info in contatos[0].get('emails', []):
                if email_info.get('tipoContato', {}).get('identificador') == 4:
                    resultado['email_funcional'] = email_info.get('endereco')
    except Exception as e:
        return "NÃO", f"Erro na consulta a API do Swagger: {str(e)}"

    return resultado

def pesquisa(item_pesquisa):
    mail = item_pesquisa.strip()
    dsquery_cmd = ["dsquery", "*", "-filter", f"(mail={mail})", "-attr", "employeeNumber"]
    try:
        cmd = subprocess.run(dsquery_cmd, capture_output=True, text=True)
        match = re.search(r'employeeNumber\s+(\d+)', cmd.stdout)
        return match.group(1) if match else None
    except Exception as e:
        return "NÃO", f"Erro na consulta por DNSQuery: {str(e)}"
        

# processa a planilha

wb = load_workbook(arquivo_excel)
ws = wb.active

log_BuscarNome = os.path.join(script_dir, "Logs_BuscarNome.txt")

# contadores de deu certo e não deu certo
success = 0
fail = 0

# lista dos erros
erros_execucao = []

print(f"🚀 Iniciando testes de buscar informações usando '{arquivo_excel}'\n")

for row in range(2, ws.max_row + 1):

    valor = ws[f'C{row}'].value
    valor = str(valor).strip() if valor else ""

    valor2 = ws[f'A{row}'].value
    valor2 = str(valor2).strip() if valor2 else ""

    if not valor and not valor2:
        ws[f'C{row}'] = "dados ausentes"
        ws[f'A{row}'] = "dados ausentes"
        ws[f'G{row}'] = "dados ausentes"
        ws[f'H{row}'] = "dados ausentes"
        print(f"Linha {row}:❌ dados ausentes")
        erros_execucao.append((row, "Dados ausentes nas colunas A e C"))

        fail += 1
        continue

    cpf = None

    if valor:
        if valor.isdigit():
            if len(valor) == 11:
                cpf = valor
            elif len(valor) == 6:
                try:
                    cpf = buscar_cpf_por_re(valor)
                except Exception as e:
                    erro = f"Erro ao buscar CPF pelo RE '{valor}': {e}"
                    erros_execucao.append((row, erro))
                    fail += 1
                    continue

    if not cpf and valor2:
        try:
            cpf = pesquisa(valor2)
        except Exception as e:
            erro = f"Erro ao pesquisar pelo email: '{valor2}'"
            print(f"Linha {row}:❌ {erro}")
            ws[f'A{row}'] = "erro ao consultar os dados pelo Email"
            erros_execucao.append((row, erro))
            fail += 1
            continue

    if cpf:
        ws[f'C{row}'] = str(cpf)
        try:
            dados = consultar_dados_por_cpf(str(cpf))
            ws[f'D{row}'] = dados.get("nome", "nome não encontrado")
            ws[f'A{row}'] = dados.get("email_funcional", "email não encontrado")
            print(f"Linha {row}:✔️ Dados encontrados pelo CPF: {cpf} | Nome: {dados.get('nome')} | Email: {dados.get('email_funcional')}")
            success += 1
        except Exception as e:
            ws[f'D{row}'] = "Nome não encontrado pelo CPF"
            ws[f'A{row}'] = "Email não encontrado pelo CPF"
            erro = f"Erro ao consultar dados pelo CPF: '{cpf}'"
            print(f"Linha {row}:❌ {erro}")
            erros_execucao.append((row, erro))
            fail += 1
    else:
        ws[f'C{row}'] = "CPF não encontrado"
        print(f"Linha {row}:❌ CPF não encontrado")
        erros_execucao.append((row, "CPF não encontrado"))
        fail += 1

# Salvar planilha
nome_arquivo = f"Credenciais_Parte1.xlsx"
resultado_path = os.path.join(script_dir, nome_arquivo)
wb.save(resultado_path)
print(f"\n📁 Arquivo salvo em: {resultado_path}\n")


# Escrever log
with open(log_BuscarNome, "a", encoding="utf-8") as log:
    log.write(f"\n=== NOVA EXECUÇÃO: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")

    log.write("-- ERROS NA EXECUÇÃO --\n")
    if erros_execucao:
        for linha, erro in erros_execucao:
            log.write(f"Linha {linha} - {erro}\n")
    else:
        log.write("Nenhum erro encontrado.\n")

    log.write(f"\n✔️ Registros com sucesso: {success}\n")
    log.write(f"❌ Registros com falha: {fail}\n")
    log.write(f"📁 Planilha gerada: {resultado_path}\n")
    log.write("==============================================\n")    


print("\n✅ Script 1 finalizado. Planilha atualizada com CPF, Nome e E-mail Funcional. \n\n")


# se tudo der certo, aqui continuar com o script 2

log_TestaCredencial = "Logs_testarCredenciais.txt"
arquivo_excel2 = os.path.join(script_dir, 'Credenciais_parte1.xlsx')

# urls para login
LOGIN_URL_MS = "http://www.sistemas.ms.com.br/login.aspx"
LOGIN_URL_AD = "http://sistemas.ad.com.br"

# fun para iniciar o driver do Selenium
def iniciar_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=options)

# valida login no MS
def validar_login_ms(driver, cpf, senha):
    try:
        driver.get(LOGIN_URL_MS)
        driver.find_element(By.NAME, "vUSRNUMCPFAUX").send_keys(cpf)
        driver.find_element(By.NAME, "vSENHA").send_keys(senha)
        driver.find_element(By.NAME, "BTN_LOGIN").click()
        time.sleep(3)
        if len(driver.window_handles) > 1 or "home.aspx" in driver.current_url.lower():
            return "SIM", "Login bem sucedido!"
        try:
            erro = driver.find_element(By.CSS_SELECTOR, "span#gxErrorViewer > div").text.strip()
            return "NÃO", erro
        except NoSuchElementException:
            return "NÃO", "Mensagem de erro não localizada"
    except Exception as e:
        return "NÃO", f"Erro ao tentar login MS: {str(e)}"

# valida login no AD
def validar_login_ad(driver, cpf, senha):
    try:
        driver.get(LOGIN_URL_AD)
        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "dropdown-toggle"))
        )
        dropdown.click()
        cpf_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "CPF"))
        )
        senha_input = driver.find_element(By.CLASS_NAME, "password")
        botao_login = driver.find_element(By.ID, "btnAutenticar")

        cpf_input.clear()
        senha_input.clear()
        cpf_input.send_keys(cpf)
        senha_input.send_keys(senha)
        botao_login.click()
        time.sleep(3)
        if "Portal/Portal?" in driver.current_url:
            return "SIM", "Login bem sucedido!"
        try:
            erro = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.ID, "msgAutenticacao"))
            ).text.strip()
            return "NÃO", erro or "Erro sem mensagem visível"
        except TimeoutException:
            return "NÃO", "Erro visível não encontrado"
    except Exception as e:
        return "NÃO", f"Erro ao tentar login AD: {str(e)}"

# executa a planilha

success_fill = PatternFill("solid", fgColor="C6EFCE")  # Verde
failure_fill = PatternFill("solid", fgColor="FFC7CE")  # Vermelho

success_ms = 0
fail_ms = 0
success_ad = 0
fail_ad = 0
erros_ms = []
erros_ad = []

wb = load_workbook(arquivo_excel2)
ws = wb.active

print(f"🚀 Iniciando testes de login usando '{arquivo_excel2}'\n")

for i in range(2, ws.max_row + 1):

    cpf = str(ws[f'C{i}'].value).strip() if ws[f'C{i}'].value else ""
    senha = str(ws[f'B{i}'].value).strip() if ws[f'B{i}'].value else ""

    if (not cpf or
        cpf.lower().startswith("cpf não encontrado") or 
        cpf.lower().startswith("dados ausentes") or 
        not senha):

        ws[f'E{i}'] = "NÃO"
        ws[f'F{i}'] = "NÃO"
        ws[f'G{i}'] = "dados ausentes"
        ws[f'H{i}'] = "dados ausentes"
        fail_ms += 1
        fail_ad += 1
        continue

    print(f"🔐 Linha {i} - Testando login | CPF: {cpf}")

    driver = None
    try:
        driver = iniciar_driver()
        status_ms, mensagem_ms = validar_login_ms(driver, cpf, senha)
        ws[f'E{i}'] = status_ms
        ws[f'G{i}'] = mensagem_ms

        if status_ms == "SIM":
            success_ms += 1
            print(f"MS = ✔️ {status_ms}")
        else:
            fail_ms += 1
            erros_ms.append((i, mensagem_ms))
            print(f"MS = ❌ {status_ms}")

        status_ad, mensagem_ad = validar_login_ad(driver, cpf, senha)
        ws[f'F{i}'] = status_ad
        ws[f'H{i}'] = mensagem_ad

        if status_ad == "SIM":
            success_ad += 1
            print(f"AD = ✔️ {status_ad}")
        else:
            fail_ad += 1
            erros_ad.append((i, mensagem_ad))
            print(f"AD = ❌ {status_ad}")

    except Exception as e:
        erro_msg = f"Erro crítico: {str(e)}"
        ws[f'E{i}'] = "NÃO"
        ws[f'F{i}'] = "NÃO"
        ws[f'G{i}'] = erro_msg
        ws[f'H{i}'] = erro_msg
        fail_ms += 1
        fail_ad += 1
        erros_ms.append((i, erro_msg))
        erros_ad.append((i, erro_msg))
    finally:
        if driver:
            driver.quit()

nome_arquivo2 = f"Credenciais_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
resultado_final = os.path.join(script_dir, nome_arquivo2)

# estilo da planilha
for i in range(2, ws.max_row + 1):
    for col in ['E', 'F']:
        cell = ws[f'{col}{i}']
        if cell.value == "SIM":
            cell.fill = success_fill
        elif cell.value == "NÃO":
            cell.fill = failure_fill
        cell.alignment = Alignment(horizontal="center")

print(f"\n✅ Testes de login concluídos. Resultados salvos em '{resultado_final}'")
print(f"\n📊 Resumo dos testes de login:")
print(f"✔️ MS - Sucessos: {success_ms} | Falhas: {fail_ms}")
print(f"✔️ AD - Sucessos: {success_ad} | Falhas: {fail_ad}")

wb.save(resultado_final)

with open(log_TestaCredencial, "a", encoding="utf-8") as log:
    log.write(f"\n=== NOVA EXECUÇÃO: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n\n")

    log.write("-- ERROS NO SISTEMA MS --\n")
    for linha, erro in erros_ms:
        log.write(f"Linha {linha} - {erro}\n")

    log.write("\n-- ERROS NO SISTEMA AD --\n")
    for linha, erro in erros_ad:
        log.write(f"Linha {linha} - {erro}\n")

    log.write(f"\n✔️ Registros MS com sucesso: {success_ms}\n")
    log.write(f"❌ Registros MS com falha: {fail_ms}\n")
    log.write(f"✔️ Registros AD com sucesso: {success_ad}\n")
    log.write(f"❌ Registros AD com falha: {fail_ad}\n")
    log.write(f"📁 Planilha gerada: {resultado_final}\n")
    log.write("==============================================\n")  


# Apaga a planilha intermediária gerada pelo Script 1
try:
    planilha_temporaria = os.path.join(script_dir, "Credenciais_Parte1.xlsx")
    if os.path.exists(planilha_temporaria):
        os.remove(planilha_temporaria)
        print(f"\n🗑️ Arquivo temporário '{planilha_temporaria}' removido com sucesso.")
except Exception as e:
    print(f"⚠️ Erro ao tentar remover o arquivo 'Credenciais_Parte1.xlsx': {e}")
